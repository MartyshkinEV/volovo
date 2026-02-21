from __future__ import annotations

import math
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from fastapi import Body, FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles

# Пытаемся импортировать модуль работы с Postgres; если его нет — используем безопасный заглушечный вариант
try:
    import pgdb as _pgdb  # type: ignore
    pgdb = _pgdb
except Exception:
    class _PgdbStub:
        def __init__(self):
            self._forms: Dict[int, Dict[str, Any]] = {}
            self._auto_id = 1

        def fetch_points(self, oid: int, d1: Optional[datetime], d2: Optional[datetime], limit: int = 500_000):
            return []

        def insert_form(self, oid: int, dt_from: Optional[datetime], dt_to: Optional[datetime], payload: Dict[str, Any], mongo_id: Optional[str] = None) -> int:
            fid = self._auto_id
            self._auto_id += 1
            self._forms[fid] = {
                "id": fid,
                "created_at": datetime.utcnow(),
                "payload": payload,
            }
            return fid

        def get_form(self, form_id: int) -> Optional[Dict[str, Any]]:
            return self._forms.get(form_id)

        def list_forms(self, limit: int = 50) -> List[Dict[str, Any]]:
            out = []
            for fid in sorted(self._forms.keys(), reverse=True)[:limit]:
                doc = self._forms[fid]
                payload = doc.get("payload") or {}
                out.append({
                    "form_id": str(fid),
                    "created_at": doc.get("created_at"),
                    "meta": payload.get("meta") or {},
                })
            return out

        def fetch_oids(self, limit: int = 500) -> List[int]:
            return []

        def fetch_routes(self) -> List[Dict[str, Any]]:
            return []

    pgdb = _PgdbStub()

# -----------------------------
# Настройки
# -----------------------------
BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_XLSX = BASE_DIR / "Камаз-маз.xlsx"

# ---- Пескобаза (погрузка) ----
SAND_BASE_LAT = 52.036242
SAND_BASE_LON = 37.887744
SAND_BASE_RADIUS_KM = 0.02  # 20 метров (0.02 км)

# -----------------------------
# App
# -----------------------------
app = FastAPI(title="Volovo Putevoy + Map + Trips + Postgres")

# статические файлы бери из проекта, а не абсолютным /opt/...
_putevoy_dir = BASE_DIR / "putevoy"
if _putevoy_dir.exists():
    app.mount("/putevoy", StaticFiles(directory=str(_putevoy_dir), html=True), name="putevoy")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# -----------------------------
# Helpers
# -----------------------------
def to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        try:
            return float(str(v).replace(",", "."))
        except Exception:
            return None


def parse_tm(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    ss = str(s).strip().replace("Z", "")
    if not ss:
        return None
    try:
        return datetime.fromisoformat(ss.replace(" ", "T"))
    except Exception:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M"):
        try:
            return datetime.strptime(ss, fmt)
        except Exception:
            continue
    return None


def fmt_tm(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(math.radians(lat1))
        * math.cos(math.radians(lat2))
        * math.sin(dlon / 2) ** 2
    )
    return 2 * R * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def iter_points_for_oid(
    oid: int,
    dt_from: Optional[str],
    dt_to: Optional[str],
    limit: int = 500_000,
) -> List[Dict[str, Any]]:
    d1 = parse_tm(dt_from) if dt_from else None
    d2 = parse_tm(dt_to) if dt_to else None

    rows = pgdb.fetch_points(oid, d1, d2, limit=limit)

    pts: List[Dict[str, Any]] = []
    for tm_dt, lon, lat, idx in rows:
        if lat is None or lon is None:
            continue
        pts.append(
            {
                "lat": float(lat),
                "lon": float(lon),
                "tm": fmt_tm(tm_dt),
                "tm_dt": tm_dt,
                "idx": idx,
                "dst": None,
                "speed": None,
            }
        )
    return pts


def calc_total_km(points: List[Dict[str, Any]]) -> float:
    if len(points) < 2:
        return 0.0
    s = 0.0
    p0 = points[0]
    for p in points[1:]:
        s += haversine_km(p0["lat"], p0["lon"], p["lat"], p["lon"])
        p0 = p
    return s


def calc_total_km_dst(points: List[Dict[str, Any]]) -> float:
    return calc_total_km(points)


def gps_filter_jumps(points: List[Dict[str, Any]], max_jump_km: float = 1.0, max_speed_kmh: float = 180.0):
    n = len(points)
    if n < 2:
        return points, {"original": n, "kept": n, "removed": 0}

    kept = [points[0]]
    removed = 0
    prev = points[0]

    for p in points[1:]:
        d = haversine_km(prev["lat"], prev["lon"], p["lat"], p["lon"])

        speed_ok = True
        t1 = prev.get("tm_dt")
        t2 = p.get("tm_dt")
        if t1 and t2:
            dt_s = (t2 - t1).total_seconds()
            if dt_s > 0:
                sp = d / (dt_s / 3600.0)
                if sp > max_speed_kmh:
                    speed_ok = False

        if d <= max_jump_km and speed_ok:
            kept.append(p)
            prev = p
        else:
            removed += 1

    return kept, {"original": n, "kept": len(kept), "removed": removed}


def count_sand_base_entries(points: List[Dict[str, Any]]) -> int:
    inside_prev = False
    entries = 0
    for p in points:
        d = haversine_km(SAND_BASE_LAT, SAND_BASE_LON, p["lat"], p["lon"])
        inside = d <= SAND_BASE_RADIUS_KM
        if inside and not inside_prev:
            entries += 1
        inside_prev = inside
    return entries


def split_trips_from_sand_base(points: List[Dict[str, Any]]):
    trips: List[List[Dict[str, Any]]] = []
    cur: List[Dict[str, Any]] = []
    inside_prev = False
    entry_idxs = []

    for i, p in enumerate(points):
        d = haversine_km(SAND_BASE_LAT, SAND_BASE_LON, p["lat"], p["lon"])
        inside = d <= SAND_BASE_RADIUS_KM

        if inside and not inside_prev:
            entry_idxs.append(i)
            if cur:
                trips.append(cur)
                cur = []
        cur.append(p)
        inside_prev = inside

    if cur:
        trips.append(cur)

    return trips, entry_idxs


def slim_points(points: List[Dict[str, Any]], max_points: int = 4000):
    n = len(points)
    if n <= max_points:
        return points, 1
    step = max(1, n // max_points)
    return points[::step], step


def _sanitize_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    meta = payload.get("meta") or {}
    rows = payload.get("rows") or []
    totals = payload.get("totals") or {}

    meta_out = {
        "oid": str(meta.get("oid") or ""),
        "dt_from": meta.get("dt_from") or "",
        "dt_to": meta.get("dt_to") or "",
    }

    rows_out = []
    for r in rows:
        rr = r or {}
        rows_out.append(
            {
                "route": rr.get("route") or "",
                "tripNo": rr.get("tripNo") or "",
                "km": rr.get("km") or "",
                "tons": rr.get("tons") or "",
                "width": rr.get("width") or "",
                "length": rr.get("length") or "",
                "pssTonnage": rr.get("pssTonnage") or "",
                "delivery": rr.get("delivery") or "",
            }
        )

    totals_out = {
        "km_spread": totals.get("km_spread") or "",
        "tons_sum": totals.get("tons_sum") or "",
        "km_gps": totals.get("km_gps") or "",
        "delivery": totals.get("delivery") or "",
        "idle": totals.get("idle") or "",
    }

    return {"meta": meta_out, "rows": rows_out, "totals": totals_out}


def fill_putevoy_xlsx(payload: Dict[str, Any], template_path: Path) -> str:
    if template_path.exists():
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()
    ws = wb.active

    totals = payload.get("totals") or {}
    ws["AF17"] = totals.get("km_gps") or ""
    ws["AF18"] = totals.get("km_spread") or ""
    ws["AF19"] = totals.get("tons_sum") or ""
    ws["AF20"] = totals.get("delivery") or ""
    ws["AF21"] = totals.get("idle") or ""

    out = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    out.close()
    wb.save(out.name)
    return out.name


# -----------------------------
# API: Forms (Postgres)
# -----------------------------
@app.post("/api/forms/save")
def save_form(payload: Dict[str, Any] = Body(...)):
    doc = _sanitize_payload(payload)

    meta = doc.get("meta") or {}
    oid = int(meta.get("oid") or 0) if str(meta.get("oid") or "").isdigit() else 0
    dt_from = parse_tm(meta.get("dt_from")) if meta.get("dt_from") else None
    dt_to = parse_tm(meta.get("dt_to")) if meta.get("dt_to") else None

    new_id = pgdb.insert_form(oid=oid, dt_from=dt_from, dt_to=dt_to, payload=doc, mongo_id=None)
    return {"status": "ok", "form_id": str(new_id)}


@app.get("/api/forms/{form_id}")
def get_form(form_id: str):
    if not form_id.isdigit():
        return JSONResponse({"error": "bad form_id"}, status_code=400)

    doc = pgdb.get_form(int(form_id))
    if not doc:
        return JSONResponse({"error": "not found"}, status_code=404)

    payload = doc.get("payload") or {}
    return {
        "form_id": str(doc.get("id") or form_id),
        "created_at": (doc.get("created_at").isoformat(timespec="seconds") if doc.get("created_at") else ""),
        **payload,
    }


@app.get("/api/forms")
def list_forms(limit: int = Query(50, ge=1, le=500)):
    forms = pgdb.list_forms(limit=limit)
    out = []
    for f in forms:
        out.append(
            {
                "form_id": f.get("form_id"),
                "created_at": f.get("created_at").isoformat(timespec="seconds") if f.get("created_at") else "",
                "meta": f.get("meta") or {},
            }
        )
    return {"forms": out}


@app.get("/api/forms/{form_id}/export_xlsx")
def export_xlsx_from_db(form_id: str):
    if not form_id.isdigit():
        return JSONResponse({"error": "bad form_id"}, status_code=400)

    doc = pgdb.get_form(int(form_id))
    if not doc:
        return JSONResponse({"error": "not found"}, status_code=404)

    payload = doc.get("payload") or {}
    out_path = fill_putevoy_xlsx(payload, TEMPLATE_XLSX)
    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"putevoy-{form_id}.xlsx",
    )


# -----------------------------
# API: OIDs list (Postgres)
# -----------------------------
@app.get("/api/oids")
def list_oids(limit: int = Query(500, ge=1, le=5000)):
    return {"oids": pgdb.fetch_oids(limit=limit)}


# -----------------------------
# API: Summary & Trips
# -----------------------------
@app.get("/api/points_summary")
def points_summary(
    oid: int = Query(...),
    dt_from: Optional[str] = Query(None),
    dt_to: Optional[str] = Query(None),
    max_jump_km: float = Query(1.0, ge=0.0, le=50.0),
    max_speed_kmh: float = Query(180.0, ge=1.0, le=400.0),
):
    pts = iter_points_for_oid(oid, dt_from, dt_to)

    km_dst = calc_total_km_dst(pts)
    pts_f, jump_stats = gps_filter_jumps(pts, max_jump_km=max_jump_km, max_speed_kmh=max_speed_kmh)
    km_hav = calc_total_km(pts_f)

    entries = count_sand_base_entries(pts_f)
    trips, entry_idxs = split_trips_from_sand_base(pts_f)
    trips_km_filtered = [tr for tr in trips if calc_total_km(tr) >= 1.0]

    return {
        "oid": oid,
        "tm_from": dt_from,
        "tm_to": dt_to,
        "points_cnt": len(pts),
        "points_cnt_filtered": len(pts_f),
        "jump_filter": jump_stats,
        "km_dst": round(km_dst, 3),
        "km_haversine": round(km_hav, 3),
        "sand_base_entries": entries,
        "trips_total": len(trips),
        "trips_filtered": len(trips_km_filtered),
    }


@app.get("/api/trips_for_map")
def trips_for_map(
    oid: int = Query(...),
    dt_from: Optional[str] = Query(None),
    dt_to: Optional[str] = Query(None),
    max_points: int = Query(4000, ge=200, le=20000),
):
    pts = iter_points_for_oid(oid, dt_from, dt_to)
    pts_f, jump_stats = gps_filter_jumps(pts)
    trips, entry_idxs = split_trips_from_sand_base(pts_f)

    out_trips = []
    for i, tr in enumerate(trips):
        tr_slim, step = slim_points(tr, max_points=max_points)
        dist_hav = calc_total_km(tr)
        dist_dst = calc_total_km_dst(tr)
        out_trips.append(
            {
                "i": i + 1,
                "points": [{"lat": p["lat"], "lon": p["lon"], "tm": p["tm"]} for p in tr_slim],
                "points_cnt": len(tr),
                "slim_step": step,
                "km_haversine": round(dist_hav, 3),
                "km_dst": round(dist_dst, 3),
                "tm_from": tr[0].get("tm") if tr else None,
                "tm_to": tr[-1].get("tm") if tr else None,
            }
        )

    return {
        "oid": oid,
        "tm_from": dt_from,
        "tm_to": dt_to,
        "points_cnt": len(pts),
        "points_cnt_filtered": len(pts_f),
        "jump_filter": jump_stats,
        "trips": out_trips,
    }


# -----------------------------
# Pages + Routes catalog (Postgres)
# -----------------------------
@app.get("/")
def home():
    return RedirectResponse(url="/putevoy/", status_code=302)


@app.get("/api/routes")
def api_routes():
    return {"routes": pgdb.fetch_routes()}