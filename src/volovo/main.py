from __future__ import annotations

import math
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from bson import ObjectId
from fastapi import Body, FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pymongo import ASCENDING, MongoClient

# -----------------------------
# Настройки
# -----------------------------
BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_XLSX = BASE_DIR / "Камаз-маз.xlsx"

MONGO_URI = "mongodb://127.0.0.1:27017"
DB_NAME = "volovo"

COL_POINTS = "track_points"        # точки
COL_ROUTES = "routes_catalog"      # справочник дорог/маршрутов
COL_FORMS = "putevoy_forms"        # сохранённые путевые листы (формы)

# ---- Пескобаза (погрузка) ----
SAND_BASE_LAT = 52.036242
SAND_BASE_LON = 37.887744
SAND_BASE_RADIUS_KM = 0.02    # 20 метров (0.02 км)

# -----------------------------
# App
# -----------------------------
app = FastAPI(title="Volovo Putevoy + Map + Trips + Mongo Forms")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

client = MongoClient(MONGO_URI)
db = client[DB_NAME]
points_col = db[COL_POINTS]
routes_col = db[COL_ROUTES]
forms_col = db[COL_FORMS]

# индексы для форм
forms_col.create_index([("created_at", ASCENDING)])
forms_col.create_index([("meta.oid", ASCENDING), ("created_at", ASCENDING)])

# -----------------------------
# Helpers
# -----------------------------
def to_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        try:
            return float(str(v).replace(",", "."))
        except Exception:
            return None


def parse_tm(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except Exception:
        return None


def build_time_query(dt_from: Optional[datetime], dt_to: Optional[datetime]) -> Dict[str, Any]:
    q: Dict[str, Any] = {}
    if dt_from or dt_to:
        q["tm"] = {}
        if dt_from:
            q["tm"]["$gte"] = dt_from.strftime("%Y-%m-%d %H:%M:%S")
        if dt_to:
            q["tm"]["$lte"] = dt_to.strftime("%Y-%m-%d %H:%M:%S")
    return q


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2) ** 2
    )
    return 2 * R * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def iter_points_for_oid(
    oid: int,
    dt_from: Optional[datetime],
    dt_to: Optional[datetime],
    limit: int = 500_000,
) -> List[Dict[str, Any]]:
    q: Dict[str, Any] = {"oid": oid}
    q.update(build_time_query(dt_from, dt_to))

    cur = (
        points_col.find(q, {"_id": 0, "lat": 1, "lon": 1, "tm": 1, "idx": 1})
        .sort([("idx", ASCENDING), ("tm", ASCENDING)])
        .limit(limit)
    )

    pts: List[Dict[str, Any]] = []
    for p in cur:
        lat = to_float(p.get("lat"))
        lon = to_float(p.get("lon"))
        if lat is None or lon is None:
            continue
        pts.append(
            {
                "lat": float(lat),
                "lon": float(lon),
                "tm": p.get("tm"),
                "tm_dt": parse_tm(p.get("tm")),
                "idx": p.get("idx"),
            }
        )
    return pts


def gps_filter_jumps(
    points: List[Dict[str, Any]],
    max_jump_km: float = 1.0,
    max_speed_kmh: float = 180.0,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    n = len(points)
    if n < 2:
        return points, {"original": n, "kept": n, "removed": 0}

    kept = [points[0]]
    removed = 0
    prev = points[0]

    for p in points[1:]:
        d = haversine_km(prev["lat"], prev["lon"], p["lat"], p["lon"])

        speed_ok = True
        t1 = prev.get("tm_dt")
        t2 = p.get("tm_dt")
        if t1 and t2:
            dt_s = (t2 - t1).total_seconds()
            if dt_s > 0:
                sp = d / (dt_s / 3600.0)  # км/ч
                if sp > max_speed_kmh:
                    speed_ok = False

        if d > max_jump_km or not speed_ok:
            removed += 1
            continue

        kept.append(p)
        prev = p

    return kept, {"original": n, "kept": len(kept), "removed": removed}


def calc_total_km(points: List[Dict[str, Any]]) -> float:
    if len(points) < 2:
        return 0.0
    total = 0.0
    prev = points[0]
    for p in points[1:]:
        total += haversine_km(prev["lat"], prev["lon"], p["lat"], p["lon"])
        prev = p
    return total


def count_sand_base_entries(points: List[Dict[str, Any]]) -> int:
    inside = False
    entries = 0

    for p in points:
        d = haversine_km(p["lat"], p["lon"], SAND_BASE_LAT, SAND_BASE_LON)
        if d <= SAND_BASE_RADIUS_KM:
            if not inside:
                entries += 1
                inside = True
        else:
            inside = False

    return entries


def split_trips_from_sand_base(points: List[Dict[str, Any]]) -> Tuple[List[List[Dict[str, Any]]], List[int]]:
    """
    РАЗБИВАЕТ НА РЕЙСЫ:
    Рейс начинается с момента ВЪЕЗДА на пескобазу (outside->inside),
    и длится до следующего въезда (следующего outside->inside).
    """
    n = len(points)
    if n == 0:
        return [], []

    inside_prev = False
    entry_indexes: List[int] = []

    for i, p in enumerate(points):
        d = haversine_km(p["lat"], p["lon"], SAND_BASE_LAT, SAND_BASE_LON)
        inside = d <= SAND_BASE_RADIUS_KM
        if inside and not inside_prev:
            entry_indexes.append(i)
        inside_prev = inside

    if not entry_indexes:
        return [points], []

    trips: List[List[Dict[str, Any]]] = []
    for k, start_i in enumerate(entry_indexes):
        end_i = entry_indexes[k + 1] if k + 1 < len(entry_indexes) else n
        seg = points[start_i:end_i]
        if len(seg) >= 2:
            trips.append(seg)

    return trips, entry_indexes


def slim_points(points: List[Dict[str, Any]], max_points: int) -> Tuple[List[Dict[str, Any]], int]:
    n = len(points)
    if n <= max_points:
        return points, 1
    step = max(1, n // max_points)
    out = points[::step]
    if out and out[-1] != points[-1]:
        out.append(points[-1])
    return out, step


# -----------------------------
# Export helpers (Excel)
# -----------------------------
def _num(v: Any) -> float:
    """
    Преобразует строку вида "184,93" / "518 т" / " 1 234,50 " в float.
    Если пусто/None/ошибка -> 0.0
    """
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0

    # оставим только цифры, минус, запятую, точку
    cleaned = "".join(ch for ch in s if ch.isdigit() or ch in ",.-")
    cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except Exception:
        return 0.0


def _set_cell(ws, row: int, col: int, value: Any):
    ws.cell(row=row, column=col).value = value


def fill_putevoy_xlsx(payload: Dict[str, Any], template_path: Path) -> Path:
    """
    ✅ Заполняем Excel так:
      - В строках 56..63:
          26 route
          27 tripNo
          29 km
          30 tons
          36 delivery (ПО СТРОКЕ)
      - Итоги в строке 64:
          29 km_spread
          30 tons_sum
          32 km_gps
          36 delivery_total
          38 idle_total
    """
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Map "column numbers" 26..40 -> actual excel columns based on row 55 markers
    col_map: Dict[int, int] = {}
    header_row = 55
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip().isdigit():
            col_map[int(v.strip())] = c

    def col(n: int) -> int:
        if n not in col_map:
            raise KeyError(f"Column marker {n} not found in template (row {header_row}).")
        return col_map[n]

    data_rows = payload.get("rows") or []
    start_row = 56
    max_lines = 8

    for i in range(max_lines):
        r = start_row + i
        item = data_rows[i] if i < len(data_rows) else {}

        route = (item.get("route") or "").strip()
        tripNo = item.get("tripNo") or str(i + 1)
        km = item.get("km") or ""
        tons = item.get("tons") or ""
        delivery = item.get("delivery") or ""

        _set_cell(ws, r, col(26), route if route else None)
        _set_cell(ws, r, col(27), tripNo if str(tripNo).strip() else None)

        # 28 не трогаем
        _set_cell(ws, r, col(29), _num(km) if str(km).strip() else None)
        _set_cell(ws, r, col(30), _num(tons) if str(tons).strip() else None)

        # ✅ 36 доставка по строке
        _set_cell(ws, r, col(36), _num(delivery) if str(delivery).strip() else None)

        # 38 (холостой) в строках не заполняем

    totals = payload.get("totals") or {}
    total_row = 64

    _set_cell(ws, total_row, col(29), _num(totals.get("km_spread")) if str(totals.get("km_spread") or "").strip() else None)
    _set_cell(ws, total_row, col(30), _num(totals.get("tons_sum")) if str(totals.get("tons_sum") or "").strip() else None)
    _set_cell(ws, total_row, col(32), _num(totals.get("km_gps")) if str(totals.get("km_gps") or "").strip() else None)

    _set_cell(ws, total_row, col(36), _num(totals.get("delivery")) if str(totals.get("delivery") or "").strip() else None)
    _set_cell(ws, total_row, col(38), _num(totals.get("idle")) if str(totals.get("idle") or "").strip() else None)

    tmpdir = Path(tempfile.gettempdir())
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    out_path = tmpdir / f"putevoy-{ts}.xlsx"
    wb.save(out_path)
    return out_path


# -----------------------------
# Forms (MongoDB)
# -----------------------------
def _sanitize_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Храним только нужное:
      meta: oid, dt_from, dt_to
      rows: route, tripNo, km, tons, width, length, pssTonnage, delivery
      totals: km_spread, tons_sum, km_gps, delivery, idle
    """
    meta = payload.get("meta") or {}
    rows = payload.get("rows") or []
    totals = payload.get("totals") or {}

    meta_out = {
        "oid": meta.get("oid") or "",
        "dt_from": meta.get("dt_from") or "",
        "dt_to": meta.get("dt_to") or "",
    }

    rows_out = []
    for r in rows:
        rr = r or {}
        rows_out.append(
            {
                "route": rr.get("route") or "",
                "tripNo": rr.get("tripNo") or "",
                "km": rr.get("km") or "",
                "tons": rr.get("tons") or "",
                "width": rr.get("width") or "",
                "length": rr.get("length") or "",
                "pssTonnage": rr.get("pssTonnage") or "",
                "delivery": rr.get("delivery") or "",  # ✅ по строке
                # idle по строкам не используем, но если прилетит — не храним, чтобы не путаться
            }
        )

    totals_out = {
        "km_spread": totals.get("km_spread") or "",
        "tons_sum": totals.get("tons_sum") or "",
        "km_gps": totals.get("km_gps") or "",
        "delivery": totals.get("delivery") or "",
        "idle": totals.get("idle") or "",
    }

    return {"meta": meta_out, "rows": rows_out, "totals": totals_out}


@app.post("/api/forms/save")
def save_form(payload: Dict[str, Any] = Body(...)):
    doc = _sanitize_payload(payload)
    doc["created_at"] = datetime.now()
    res = forms_col.insert_one(doc)
    return {"status": "ok", "form_id": str(res.inserted_id)}


@app.get("/api/forms/{form_id}")
def get_form(form_id: str):
    try:
        oid = ObjectId(form_id)
    except Exception:
        return JSONResponse({"error": "bad form_id"}, status_code=400)

    doc = forms_col.find_one({"_id": oid})
    if not doc:
        return JSONResponse({"error": "not found"}, status_code=404)

    doc["_id"] = str(doc["_id"])
    if isinstance(doc.get("created_at"), datetime):
        doc["created_at"] = doc["created_at"].isoformat(timespec="seconds")
    return doc


@app.get("/api/forms")
def list_forms(limit: int = Query(50, ge=1, le=500)):
    cur = forms_col.find({}, {"meta": 1, "created_at": 1}).sort([("created_at", -1)]).limit(limit)
    out = []
    for d in cur:
        out.append(
            {
                "form_id": str(d["_id"]),
                "created_at": d.get("created_at").isoformat(timespec="seconds")
                if isinstance(d.get("created_at"), datetime)
                else "",
                "meta": d.get("meta") or {},
            }
        )
    return {"forms": out}


@app.get("/api/forms/{form_id}/export_xlsx")
def export_xlsx_from_db(form_id: str):
    try:
        oid = ObjectId(form_id)
    except Exception:
        return JSONResponse({"error": "bad form_id"}, status_code=400)

    doc = forms_col.find_one({"_id": oid})
    if not doc:
        return JSONResponse({"error": "not found"}, status_code=404)

    payload = {
        "rows": doc.get("rows") or [],
        "totals": doc.get("totals") or {},
    }

    out_path = fill_putevoy_xlsx(payload, TEMPLATE_XLSX)
    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"putevoy-{form_id}.xlsx",
    )


# -----------------------------
# Pages
# -----------------------------
@app.get("/")
def home():
    return FileResponse(BASE_DIR / "putevoy.html")


@app.get("/putevoy")
def putevoy_page():
    return FileResponse(BASE_DIR / "putevoy.html")


# -----------------------------
# API: routes catalog
# -----------------------------
@app.get("/api/routes")
def api_routes(limit: int = Query(5000, ge=1, le=50000)):
    cur = (
        routes_col.find(
            {},
            {"_id": 0, "name": 1, "road_width_m": 1, "pss_tonnage_t": 1, "road_length_km": 1},
        )
        .sort([("name", ASCENDING)])
        .limit(limit)
    )
    return {"routes": list(cur)}


@app.get("/api/oids")
def list_oids(limit: int = Query(2000, ge=1, le=20000)):
    oids = points_col.distinct("oid")
    oids = sorted([int(x) for x in oids if x is not None])[:limit]
    return {"oids": oids}


# -----------------------------
# (совместимость) API: export xlsx из payload (не из БД)
# -----------------------------
@app.post("/api/export_xlsx")
async def export_xlsx(payload: Dict[str, Any] = Body(...)):
    out_path = fill_putevoy_xlsx(payload, TEMPLATE_XLSX)
    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=out_path.name,
    )


# -----------------------------
# API: summary
# -----------------------------
@app.get("/api/points_summary")
def points_summary(
    oid: int = Query(...),
    dt_from: Optional[str] = Query(None, description="YYYY-MM-DD HH:MM:SS"),
    dt_to: Optional[str] = Query(None, description="YYYY-MM-DD HH:MM:SS"),
    limit: int = Query(500000, ge=10, le=1000000),
    max_jump_km: float = Query(1.0, ge=0.05, le=50.0),
    max_speed_kmh: float = Query(180.0, ge=10.0, le=500.0),
):
    dt_from_dt = parse_tm(dt_from)
    dt_to_dt = parse_tm(dt_to)

    pts = iter_points_for_oid(oid=oid, dt_from=dt_from_dt, dt_to=dt_to_dt, limit=limit)
    pts_f, gps_stats = gps_filter_jumps(pts, max_jump_km=max_jump_km, max_speed_kmh=max_speed_kmh)

    total_km = calc_total_km(pts_f)
    sand_base_entries = count_sand_base_entries(pts_f)

    trips, _entry_idx = split_trips_from_sand_base(pts_f)
    trips_km_filtered = [tr for tr in trips if calc_total_km(tr) >= 1.0]

    return JSONResponse(
        {
            "oid": oid,
            "dt_from": dt_from,
            "dt_to": dt_to,
            "points_count_original": len(pts),
            "points_count_used": len(pts_f),
            "gps_jumps_removed": gps_stats["removed"],
            "max_jump_km": max_jump_km,
            "max_speed_kmh": max_speed_kmh,
            "total_km": total_km,
            "sand_base_entries": sand_base_entries,
            "trips_count_raw": len(trips),
            "trips_count_ge1km": len(trips_km_filtered),
        }
    )


# -----------------------------
# API: trips for map (слои рейсов)
# -----------------------------
@app.get("/api/trips_for_map")
def trips_for_map(
    oid: int = Query(...),
    dt_from: Optional[str] = Query(None, description="YYYY-MM-DD HH:MM:SS"),
    dt_to: Optional[str] = Query(None, description="YYYY-MM-DD HH:MM:SS"),
    limit: int = Query(500000, ge=10, le=1000000),
    max_points_per_trip: int = Query(2000, ge=200, le=20000),
    max_jump_km: float = Query(1.0, ge=0.05, le=50.0),
    max_speed_kmh: float = Query(180.0, ge=10.0, le=500.0),
    min_trip_km: float = Query(1.0, ge=0.0, le=5000.0),
):
    dt_from_dt = parse_tm(dt_from)
    dt_to_dt = parse_tm(dt_to)

    pts = iter_points_for_oid(oid=oid, dt_from=dt_from_dt, dt_to=dt_to_dt, limit=limit)
    pts_f, gps_stats = gps_filter_jumps(pts, max_jump_km=max_jump_km, max_speed_kmh=max_speed_kmh)

    trips, _entry_indexes = split_trips_from_sand_base(pts_f)

    trips_out: List[Dict[str, Any]] = []
    trip_no = 0

    for tr in trips:
        dist_km = calc_total_km(tr)
        if dist_km < min_trip_km:
            continue

        trip_no += 1
        slim, step = slim_points(tr, max_points=max_points_per_trip)
        points_out = [{"lat": p["lat"], "lon": p["lon"], "tm": p.get("tm")} for p in slim]

        tm_start = tr[0].get("tm")
        tm_end = tr[-1].get("tm")

        trips_out.append(
            {
                "trip_no": trip_no,
                "tm_start": tm_start,
                "tm_end": tm_end,
                "original_points": len(tr),
                "step": step,
                "distance_km": dist_km,
                "points": points_out,
            }
        )

    return {
        "oid": oid,
        "dt_from": dt_from,
        "dt_to": dt_to,
        "original_count": len(pts),
        "filtered_count": len(pts_f),
        "gps_jumps_removed": gps_stats["removed"],
        "sand_base_entries": count_sand_base_entries(pts_f),
        "trips_count": len(trips_out),
        "min_trip_km": min_trip_km,
        "max_points_per_trip": max_points_per_trip,
        "sand_base": {
            "lat": SAND_BASE_LAT,
            "lon": SAND_BASE_LON,
            "radius_km": SAND_BASE_RADIUS_KM,
        },
        "trips": trips_out,
    }
